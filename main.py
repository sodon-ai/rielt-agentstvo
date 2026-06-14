from fastapi import FastAPI, Request, Form, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import uvicorn
import os
import uuid
import shutil
import json
from datetime import datetime, timedelta
from typing import Optional, List
import re
import smtplib
from email.mime.text import MimeText
from email.mime.multipart import MimeMultipart
import random
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

app = FastAPI(title="Риэлторское агентство by Гаун")

# Создаём папки
os.makedirs("static/uploads/photos", exist_ok=True)
os.makedirs("static/uploads/property_photos", exist_ok=True)
os.makedirs("data", exist_ok=True)

app.mount("/static", StaticFiles(directory="static"), name="static")

# Конфигурация почты (для уведомлений и восстановления пароля)
EMAIL_CONFIG = {
    "smtp_server": "smtp.gmail.com",
    "smtp_port": 587,
    "sender_email": "artemgaun104@gmail.com",
    "sender_password": "YOUR_APP_PASSWORD"  # Тут нужно будет вставить пароль приложения Gmail
}

# Файлы для хранения данных
DATA_FILES = {
    "users": "data/users.json",
    "properties": "data/properties.json",
    "messages": "data/messages.json",
    "private_messages": "data/private_messages.json",
    "notifications": "data/notifications.json",
    "transactions": "data/transactions.json",
    "deleted_messages": "data/deleted_messages.json",
    "reviews": "data/reviews.json",
    "banned_users": "data/banned_users.json",
    "favorites": "data/favorites.json",
    "views_history": "data/views_history.json",
    "comparisons": "data/comparisons.json",
    "email_verifications": "data/email_verifications.json",
    "password_resets": "data/password_resets.json"
}

def send_email(to_email, subject, html_content):
    """Отправка email через SMTP"""
    try:
        msg = MimeMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = EMAIL_CONFIG["sender_email"]
        msg["To"] = to_email
        
        html_part = MimeText(html_content, "html")
        msg.attach(html_part)
        
        server = smtplib.SMTP(EMAIL_CONFIG["smtp_server"], EMAIL_CONFIG["smtp_port"])
        server.starttls()
        server.login(EMAIL_CONFIG["sender_email"], EMAIL_CONFIG["sender_password"])
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Ошибка отправки email: {e}")
        return False

def send_verification_email(to_email, username, code):
    """Отправка письма с подтверждением регистрации"""
    html = f"""
    <html>
    <head><meta charset="UTF-8"></head>
    <body style="font-family: Arial, sans-serif;">
        <h2>Добро пожаловать в Агентство «by Гаун»!</h2>
        <p>Здравствуйте, {username}!</p>
        <p>Для завершения регистрации подтвердите ваш email, перейдя по ссылке:</p>
        <p><a href="https://rielt-agentstvo.onrender.com/verify_email?code={code}" style="background:#6ec8e6;padding:10px 20px;color:#1a4d66;text-decoration:none;border-radius:25px;">Подтвердить email</a></p>
        <p>Если вы не регистрировались на нашем сайте, просто проигнорируйте это письмо.</p>
        <hr>
        <p>С уважением,<br>Команда Агентства «by Гаун»</p>
    </body>
    </html>
    """
    return send_email(to_email, "Подтверждение регистрации", html)

def send_reset_email(to_email, username, code):
    """Отправка письма для восстановления пароля"""
    html = f"""
    <html>
    <head><meta charset="UTF-8"></head>
    <body style="font-family: Arial, sans-serif;">
        <h2>Восстановление пароля</h2>
        <p>Здравствуйте, {username}!</p>
        <p>Вы запросили восстановление пароля. Перейдите по ссылке, чтобы задать новый пароль:</p>
        <p><a href="https://rielt-agentstvo.onrender.com/reset_password_form?code={code}" style="background:#6ec8e6;padding:10px 20px;color:#1a4d66;text-decoration:none;border-radius:25px;">Сбросить пароль</a></p>
        <p>Если вы не запрашивали восстановление пароля, просто проигнорируйте это письмо.</p>
        <hr>
        <p>С уважением,<br>Команда Агентства «by Гаун»</p>
    </body>
    </html>
    """
    return send_email(to_email, "Восстановление пароля", html)

def send_notification_email(to_email, username, subject, message):
    """Отправка уведомления на почту"""
    html = f"""
    <html>
    <head><meta charset="UTF-8"></head>
    <body style="font-family: Arial, sans-serif;">
        <h2>{subject}</h2>
        <p>Здравствуйте, {username}!</p>
        <p>{message}</p>
        <hr>
        <p>С уважением,<br>Команда Агентства «by Гаун»</p>
    </body>
    </html>
    """
    return send_email(to_email, subject, html)

# Инициализация файлов
def init_data_files():
    admin_exists = False
    if os.path.exists(DATA_FILES["users"]):
        with open(DATA_FILES["users"], "r", encoding="utf-8") as f:
            users = json.load(f)
            for u in users:
                if u["username"] == "artem_gaun":
                    admin_exists = True
                    break
    
    for name, path in DATA_FILES.items():
        if not os.path.exists(path):
            with open(path, "w", encoding="utf-8") as f:
                if name == "users":
                    if not admin_exists:
                        json.dump([{
                            "id": 1,
                            "username": "artem_gaun",
                            "password": "Admin_321",
                            "nickname": "Артём Гаун",
                            "name": "Администратор",
                            "role": "admin",
                            "email": "artemgaun104@gmail.com",
                            "email_verified": True,
                            "is_active": True,
                            "is_banned": False,
                            "theme": "light",
                            "created_at": datetime.now().isoformat()
                        }], f, ensure_ascii=False, indent=2)
                    else:
                        json.dump([], f, ensure_ascii=False, indent=2)
                elif name in ["favorites", "views_history", "comparisons", "email_verifications", "password_resets", "private_messages"]:
                    json.dump([], f, ensure_ascii=False, indent=2)
                elif name == "banned_users":
                    json.dump([], f, ensure_ascii=False, indent=2)
                elif name == "reviews":
                    json.dump([], f, ensure_ascii=False, indent=2)
                else:
                    json.dump([], f, ensure_ascii=False, indent=2)

init_data_files()

def load_data(filename):
    with open(filename, "r", encoding="utf-8") as f:
        return json.load(f)

def save_data(filename, data):
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_next_id(data):
    return max([item["id"] for item in data]) + 1 if data else 1

def get_user_by_username(username):
    users = load_data(DATA_FILES["users"])
    for user in users:
        if user["username"] == username:
            return user
    return None

def get_user_by_email(email):
    users = load_data(DATA_FILES["users"])
    for user in users:
        if user.get("email") == email:
            return user
    return None

def get_user_by_nickname(nickname):
    users = load_data(DATA_FILES["users"])
    for user in users:
        if user.get("nickname") == nickname:
            return user
    return None

def update_user(user_id, updates):
    users = load_data(DATA_FILES["users"])
    for user in users:
        if user["id"] == user_id:
            user.update(updates)
            break
    save_data(DATA_FILES["users"], users)

def update_user_by_username(username, updates):
    users = load_data(DATA_FILES["users"])
    for user in users:
        if user["username"] == username:
            user.update(updates)
            break
    save_data(DATA_FILES["users"], users)

def create_user(username, password, nickname, name, role, email=""):
    users = load_data(DATA_FILES["users"])
    new_id = get_next_id(users)
    users.append({
        "id": new_id,
        "username": username,
        "password": password,
        "nickname": nickname,
        "name": name,
        "role": role,
        "email": email,
        "email_verified": False,
        "is_active": True,
        "is_banned": False,
        "theme": "light",
        "created_at": datetime.now().isoformat()
    })
    save_data(DATA_FILES["users"], users)

def get_all_users():
    return load_data(DATA_FILES["users"])

def is_user_banned(username):
    banned = load_data(DATA_FILES["banned_users"])
    for b in banned:
        if b["username"] == username:
            return True
    return False

def ban_user(username, banned_by):
    banned = load_data(DATA_FILES["banned_users"])
    if not any(b["username"] == username for b in banned):
        banned.append({
            "id": get_next_id(banned),
            "username": username,
            "banned_by": banned_by,
            "banned_at": datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        })
        save_data(DATA_FILES["banned_users"], banned)
        update_user_by_username(username, {"is_active": False, "is_banned": True})

def unban_user(username):
    banned = load_data(DATA_FILES["banned_users"])
    banned = [b for b in banned if b["username"] != username]
    save_data(DATA_FILES["banned_users"], banned)
    update_user_by_username(username, {"is_active": True, "is_banned": False})

def get_banned_users():
    return load_data(DATA_FILES["banned_users"])

def create_verification_code(user_id, email):
    codes = load_data(DATA_FILES["email_verifications"])
    code = str(random.randint(100000, 999999))
    codes.append({
        "id": get_next_id(codes),
        "user_id": user_id,
        "email": email,
        "code": code,
        "created_at": datetime.now().isoformat(),
        "expires_at": (datetime.now() + timedelta(hours=24)).isoformat()
    })
    save_data(DATA_FILES["email_verifications"], codes)
    return code

def verify_email_code(code):
    codes = load_data(DATA_FILES["email_verifications"])
    for c in codes:
        if c["code"] == code:
            if datetime.fromisoformat(c["expires_at"]) > datetime.now():
                update_user_by_username(get_user_by_username(c["email"])["username"], {"email_verified": True})
                codes.remove(c)
                save_data(DATA_FILES["email_verifications"], codes)
                return True
    return False

def create_reset_code(user_id, email):
    codes = load_data(DATA_FILES["password_resets"])
    code = str(random.randint(100000, 999999)) + str(uuid.uuid4())[:8]
    codes.append({
        "id": get_next_id(codes),
        "user_id": user_id,
        "email": email,
        "code": code,
        "created_at": datetime.now().isoformat(),
        "expires_at": (datetime.now() + timedelta(hours=1)).isoformat()
    })
    save_data(DATA_FILES["password_resets"], codes)
    return code

def verify_reset_code(code):
    codes = load_data(DATA_FILES["password_resets"])
    for c in codes:
        if c["code"] == code:
            if datetime.fromisoformat(c["expires_at"]) > datetime.now():
                return c["user_id"], c["email"]
    return None, None

def delete_reset_code(code):
    codes = load_data(DATA_FILES["password_resets"])
    codes = [c for c in codes if c["code"] != code]
    save_data(DATA_FILES["password_resets"], codes)

def get_all_properties():
    return load_data(DATA_FILES["properties"])

def add_property(title, location, price, area, floor, total_floors, building_type, description, seller_name, seller_username, photos=None):
    properties = load_data(DATA_FILES["properties"])
    new_id = get_next_id(properties)
    properties.append({
        "id": new_id,
        "title": title,
        "location": location,
        "price": price,
        "area": area,
        "floor": floor,
        "total_floors": total_floors,
        "building_type": building_type,
        "description": description,
        "seller_name": seller_name,
        "seller_username": seller_username,
        "photos": photos or [],
        "views": 0,
        "created_at": datetime.now().isoformat()
    })
    save_data(DATA_FILES["properties"], properties)
    return new_id

def update_property(prop_id, updates):
    properties = load_data(DATA_FILES["properties"])
    for prop in properties:
        if prop["id"] == prop_id:
            prop.update(updates)
            break
    save_data(DATA_FILES["properties"], properties)

def delete_property_by_id(prop_id):
    properties = load_data(DATA_FILES["properties"])
    properties = [p for p in properties if p["id"] != prop_id]
    save_data(DATA_FILES["properties"], properties)

def increment_property_views(prop_id):
    properties = load_data(DATA_FILES["properties"])
    for prop in properties:
        if prop["id"] == prop_id:
            prop["views"] = prop.get("views", 0) + 1
            break
    save_data(DATA_FILES["properties"], properties)

def get_all_messages():
    return load_data(DATA_FILES["messages"])

def add_message(username, nickname, text, time, photo):
    messages = load_data(DATA_FILES["messages"])
    new_id = get_next_id(messages)
    messages.append({
        "id": new_id,
        "username": username,
        "nickname": nickname,
        "text": text,
        "time": time,
        "photo": photo
    })
    save_data(DATA_FILES["messages"], messages)

def delete_message_by_id(msg_id, deleted_by):
    messages = load_data(DATA_FILES["messages"])
    deleted_msgs = load_data(DATA_FILES["deleted_messages"])
    msg_to_delete = None
    for msg in messages:
        if msg["id"] == msg_id:
            msg_to_delete = msg
            break
    if msg_to_delete:
        new_id = get_next_id(deleted_msgs)
        deleted_msgs.append({
            "id": new_id,
            "username": msg_to_delete["username"],
            "nickname": msg_to_delete["nickname"],
            "text": msg_to_delete["text"],
            "time": msg_to_delete["time"],
            "photo": msg_to_delete.get("photo"),
            "deleted_by": deleted_by,
            "deleted_at": datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        })
        save_data(DATA_FILES["deleted_messages"], deleted_msgs)
        messages = [m for m in messages if m["id"] != msg_id]
        save_data(DATA_FILES["messages"], messages)

def clear_deleted_messages():
    save_data(DATA_FILES["deleted_messages"], [])

def get_deleted_messages():
    return load_data(DATA_FILES["deleted_messages"])

def get_private_messages(user1, user2):
    msgs = load_data(DATA_FILES["private_messages"])
    return [m for m in msgs if (m["from_user"] == user1 and m["to_user"] == user2) or (m["from_user"] == user2 and m["to_user"] == user1)]

def add_private_message(from_user, from_nickname, to_user, text, time):
    msgs = load_data(DATA_FILES["private_messages"])
    new_id = get_next_id(msgs)
    msgs.append({
        "id": new_id,
        "from_user": from_user,
        "from_nickname": from_nickname,
        "to_user": to_user,
        "text": text,
        "time": time,
        "read": False
    })
    save_data(DATA_FILES["private_messages"], msgs)
    return new_id

def mark_private_messages_read(user1, user2):
    msgs = load_data(DATA_FILES["private_messages"])
    for msg in msgs:
        if msg["to_user"] == user1 and msg["from_user"] == user2:
            msg["read"] = True
    save_data(DATA_FILES["private_messages"], msgs)

def get_unread_private_count(username):
    msgs = load_data(DATA_FILES["private_messages"])
    return len([m for m in msgs if m["to_user"] == username and not m["read"]])

def add_notification(to_user, message, date, type, property_id):
    notifications = load_data(DATA_FILES["notifications"])
    new_id = get_next_id(notifications)
    notifications.append({
        "id": new_id,
        "to_user": to_user,
        "message": message,
        "date": date,
        "type": type,
        "property_id": property_id,
        "read": False
    })
    save_data(DATA_FILES["notifications"], notifications)
    
    # Отправляем email уведомление, если у пользователя есть почта
    user = get_user_by_username(to_user)
    if user and user.get("email") and user.get("email_verified"):
        send_notification_email(user["email"], user["nickname"], "Новое уведомление", message)

def get_notifications_by_user(username):
    notifications = load_data(DATA_FILES["notifications"])
    return [n for n in notifications if n["to_user"] == username]

def mark_notifications_read(username):
    notifications = load_data(DATA_FILES["notifications"])
    for n in notifications:
        if n["to_user"] == username:
            n["read"] = True
    save_data(DATA_FILES["notifications"], notifications)

def add_transaction(buyer, seller, property_title, price, type, date):
    transactions = load_data(DATA_FILES["transactions"])
    new_id = get_next_id(transactions)
    transactions.append({
        "id": new_id,
        "buyer": buyer,
        "seller": seller,
        "property_title": property_title,
        "price": price,
        "type": type,
        "date": date
    })
    save_data(DATA_FILES["transactions"], transactions)

def get_all_transactions():
    return load_data(DATA_FILES["transactions"])

def add_review(property_id, username, nickname, rating, comment):
    reviews = load_data(DATA_FILES["reviews"])
    new_id = get_next_id(reviews)
    reviews.append({
        "id": new_id,
        "property_id": property_id,
        "username": username,
        "nickname": nickname,
        "rating": rating,
        "comment": comment[:200],
        "date": datetime.now().strftime("%d.%m.%Y %H:%M")
    })
    save_data(DATA_FILES["reviews"], reviews)

def get_reviews_by_property(property_id):
    reviews = load_data(DATA_FILES["reviews"])
    return [r for r in reviews if r["property_id"] == property_id]

def add_favorite(user_id, property_id):
    favorites = load_data(DATA_FILES["favorites"])
    if not any(f["user_id"] == user_id and f["property_id"] == property_id for f in favorites):
        favorites.append({
            "id": get_next_id(favorites),
            "user_id": user_id,
            "property_id": property_id,
            "created_at": datetime.now().isoformat()
        })
        save_data(DATA_FILES["favorites"], favorites)

def remove_favorite(user_id, property_id):
    favorites = load_data(DATA_FILES["favorites"])
    favorites = [f for f in favorites if not (f["user_id"] == user_id and f["property_id"] == property_id)]
    save_data(DATA_FILES["favorites"], favorites)

def get_user_favorites(user_id):
    favorites = load_data(DATA_FILES["favorites"])
    return [f["property_id"] for f in favorites if f["user_id"] == user_id]

def add_view_history(user_id, property_id):
    history = load_data(DATA_FILES["views_history"])
    # Удаляем старые записи об этом объявлении
    history = [h for h in history if not (h["user_id"] == user_id and h["property_id"] == property_id)]
    history.append({
        "id": get_next_id(history),
        "user_id": user_id,
        "property_id": property_id,
        "viewed_at": datetime.now().isoformat()
    })
    # Оставляем только последние 20 просмотров
    user_history = [h for h in history if h["user_id"] == user_id]
    if len(user_history) > 20:
        to_remove = sorted(user_history, key=lambda x: x["viewed_at"])[:-20]
        history = [h for h in history if h not in to_remove]
    save_data(DATA_FILES["views_history"], history)

def get_user_view_history(user_id):
    history = load_data(DATA_FILES["views_history"])
    return [h for h in history if h["user_id"] == user_id][-20:]

def get_comparison(user_id):
    comparisons = load_data(DATA_FILES["comparisons"])
    for c in comparisons:
        if c["user_id"] == user_id:
            return c.get("properties", [])
    return []

def add_to_comparison(user_id, property_id):
    comparisons = load_data(DATA_FILES["comparisons"])
    found = False
    for c in comparisons:
        if c["user_id"] == user_id:
            if property_id not in c["properties"] and len(c["properties"]) < 3:
                c["properties"].append(property_id)
                found = True
            break
    if not found:
        comparisons.append({
            "id": get_next_id(comparisons),
            "user_id": user_id,
            "properties": [property_id]
        })
    save_data(DATA_FILES["comparisons"], comparisons)

def remove_from_comparison(user_id, property_id):
    comparisons = load_data(DATA_FILES["comparisons"])
    for c in comparisons:
        if c["user_id"] == user_id:
            if property_id in c["properties"]:
                c["properties"].remove(property_id)
            break
    save_data(DATA_FILES["comparisons"], comparisons)

def clear_comparison(user_id):
    comparisons = load_data(DATA_FILES["comparisons"])
    comparisons = [c for c in comparisons if c["user_id"] != user_id]
    save_data(DATA_FILES["comparisons"], comparisons)

def is_valid_email(email):
    if not email:
        return True
    pattern = r'^[a-zA-Z0-9_.+-]+@(mail\.ru|gmail\.com)$'
    return re.match(pattern, email) is not None

def export_transactions_to_excel(transactions):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сделки"
    headers = ["Дата", "Продавец", "Покупатель", "Объект", "Цена", "Тип"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="6ec8e6", end_color="6ec8e6", fill_type="solid")
    for t in transactions:
        ws.append([t["date"], t["seller"], t["buyer"], t["property_title"], t["price"], t["type"]])
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    return wb
    # ============= HTML ШАБЛОН =============
def render_html(content, user=None, error=None, success=None, theme=None):
    if user and user.get("theme"):
        theme = user.get("theme")
    elif not theme:
        theme = "light"
    
    user_html = ""
    if user:
        notif_list = get_notifications_by_user(user["username"])
        unread = len([n for n in notif_list if not n.get("read")])
        private_unread = get_unread_private_count(user["username"])
        total_unread = unread + private_unread
        badge = f'<span style="background:red;color:white;border-radius:50%;padding:2px 6px;font-size:12px;margin-left:5px;">{total_unread}</span>' if total_unread > 0 else ''
        user_html = f'<span>👤 {user.get("nickname", user["name"])} ({user["role"]})</span><a href="/profile" class="btn">📋 Профиль{badge}</a><a href="/logout" class="btn">🚪 Выйти</a>'
    else:
        user_html = '<a href="/login" class="btn">🔓 Вход</a><a href="/register" class="btn">📝 Регистрация</a>'
    
    err_html = f'<div class="error">{error}</div>' if error else ''
    suc_html = f'<div class="success">{success}</div>' if success else ''
    
    theme_toggle = f'<button onclick="toggleTheme()" class="btn" style="background:transparent;border:1px solid var(--light-blue);">{"🌙" if theme=="light" else "☀️"}</button>'
    
    return f'''<!DOCTYPE html>
<html lang="ru">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>Риэлторское агентство by Гаун</title>
<style>
:root {'{'}'--light-blue:#6ec8e6;'--medium-blue:#4ab0d0;'--dark-blue:#2c7aa0;'--bg:#f0f7fa;'--card-bg:#ffffff;'--text:#2c3e50;'--text-light:#4a627a;'--border:#d4e8f0;'--header-bg:#ffffff;'--shadow:0 2px 15px rgba(0,0,0,0.08);'}'
[data-theme="dark"] {'{'}--light-blue:#2c7aa0;--medium-blue:#1a5d7a;--dark-blue:#6ec8e6;--bg:#1a2a3a;--card-bg:#2a3a4a;--text:#e0e0e0;--text-light:#b0b0b0;--border:#3a4a5a;--header-bg:#1a2a3a;--shadow:0 2px 15px rgba(0,0,0,0.3);'}'}
*{'margin':0;'padding':0;'box-sizing':'border-box';'}
body{'font-family':'Segoe UI, sans-serif';'background':'var(--bg)';'color':'var(--text)';'transition':'background 0.3s, color 0.3s';}
header{'background':'var(--header-bg)';'box-shadow':'var(--shadow)';'position':'sticky';'top':0;'z-index':100;}
.top-banner{'background':'var(--light-blue)';'padding':'8px 0';'text-align':'center';'font-size':'13px';'color':'#1a4d66';}
.container{'max-width':'1200px';'margin':'0 auto';'padding':'0 25px';}
.header-main{'display':'flex';'justify-content':'space-between';'align-items':'center';'padding':'20px 0';'flex-wrap':'wrap';'gap':'15px';}
.logo-area{'display':'flex';'align-items':'center';'gap':'15px';}
.logo-icon{'background':'var(--light-blue)';'width':'55px';'height':'55px';'border-radius':'50%';'display':'flex';'align-items':'center';'justify-content':'center';'font-size':'28px';'color':'white';}
.logo-text h1{'font-size':'18px';'color':'#1a4d66';}
.logo-text h2{'font-size':'22px';'color':'var(--dark-blue)';}
.slogan{'font-size':'13px';'color':'var(--medium-blue)';}
nav ul{'display':'flex';'list-style':'none';'gap':'25px';'flex-wrap':'wrap';}
nav a{'text-decoration':'none';'color':'var(--text)';'font-weight':'600';'padding':'8px 12px';'border-radius':'8px';'transition':'0.3s';}
nav a:hover{'background':'var(--bg)';'color':'var(--dark-blue)';}
.user-info{'display':'flex';'gap':'15px';'align-items':'center';'flex-wrap':'wrap';}
.btn{'background':'var(--light-blue)';'padding':'8px 20px';'border-radius':'25px';'text-decoration':'none';'color':'#1a4d66';'font-weight':'bold';'border':'none';'cursor':'pointer';'display':'inline-block';'transition':'0.3s';}
.btn:hover{'transform':'scale(1.02)';'background':'var(--medium-blue)';}
.hero{'background':'linear-gradient(120deg, var(--bg) 0%, var(--card-bg) 80%)';'padding':'50px 0';'text-align':'center';}
.hero h1{'font-size':'42px';'color':'var(--dark-blue)';}
.services-block{'background':'var(--card-bg)';'padding':'50px 0';'border-bottom':'1px solid var(--border)';}
.services-grid{'display':'grid';'grid-template-columns':'repeat(auto-fit, minmax(320px, 1fr));'gap':'30px';'margin-top':'30px';}
.service-box{'background':'var(--bg)';'border-radius':'16px';'padding':'25px';'border-left':'4px solid var(--light-blue)';}
.service-box h3{'color':'var(--dark-blue)';'font-size':'22px';}
.service-box ul{'list-style':'none';}
.service-box li{'padding':'8px 0 8px 25px';'position':'relative';}
.service-box li:before{'content':'•';'color':'var(--light-blue)';'position':'absolute';'left':'5px';}
.about-section{'background':'var(--card-bg)';'padding':'50px 0';'border-top':'1px solid var(--border)';'border-bottom':'1px solid var(--border)';}
.about-section h2{'font-size':'32px';'color':'var(--dark-blue)';}
.about-section h3{'font-size':'26px';'color':'var(--dark-blue)';'margin':'30px 0 15px';}
.about-section p{'font-size':'16px';'line-height':'1.7';'color':'var(--text-light)';'margin-bottom':'18px';}
.about-list{'list-style':'none';}
.about-list li{'padding':'10px 0 10px 28px';'position':'relative';}
.about-list li:before{'content':'•';'color':'var(--light-blue)';'font-size':'22px';'position':'absolute';'left':'5px';}
.search-bar{'background':'var(--card-bg)';'padding':'20px';'border-radius':'16px';'margin-bottom':'30px';'display':'flex';'gap':'15px';'flex-wrap':'wrap';'align-items':'flex-end';}
.search-group{'flex':'1';'min-width':'150px';}
.search-group label{'display':'block';'margin-bottom':'5px';'font-weight':'bold';}
.search-group input,.search-group select{'width':'100%';'padding':'8px 12px';'border':'1px solid var(--border)';'border-radius':'8px';'background':'var(--bg)';'color':'var(--text)';}
.cards-grid{'display':'grid';'grid-template-columns':'repeat(auto-fill, minmax(360px, 1fr));'gap':'25px';'padding':'20px 0';}
.property-card{'background':'var(--card-bg)';'border-radius':'16px';'padding':'20px';'box-shadow':'var(--shadow)';'border-left':'4px solid var(--light-blue)';'transition':'0.2s';}
.property-card:hover{'transform':'translateY(-3px)';'box-shadow':'0 8px 20px rgba(0,0,0,0.15)';}
.property-card h3{'color':'var(--dark-blue)';'margin-bottom':'10px';}
.property-photos{'display':'flex';'gap':'8px';'margin-top':'10px';'flex-wrap':'wrap';}
.property-photo{'width':'80px';'height':'80px';'object-fit':'cover';'border-radius':'8px';'cursor':'pointer';'transition':'0.2s';}
.property-photo:hover{'transform':'scale(1.05);'}
.property-actions{'display':'flex';'gap':'10px';'margin-top':'15px';'flex-wrap':'wrap';}
.delete-btn{'background':'#ff6b6b';'color':'white';'border':'none';'padding':'8px 15px';'border-radius':'8px';'cursor':'pointer';}
.rent-btn{'background':'#4caf50';'color':'white';'border':'none';'padding':'8px 15px';'border-radius':'8px';'cursor':'pointer';}
.buy-btn{'background':'var(--light-blue)';'color':'#1a4d66';'border':'none';'padding':'8px 15px';'border-radius':'8px';'cursor':'pointer';}
.compare-btn{'background':'#ff9800';'color':'white';'border':'none';'padding':'8px 15px';'border-radius':'8px';'cursor':'pointer';}
.fav-btn{'background':'transparent';'border':'none';'font-size':'24px';'cursor':'pointer';}
.ban-btn{'background':'#ff9800';'color':'white';'border':'none';'padding':'5px 10px';'border-radius':'6px';'cursor':'pointer';'font-size':'12px';}
.unban-btn{'background':'#4caf50';'color':'white';'border':'none';'padding':'5px 10px';'border-radius':'6px';'cursor':'pointer';'font-size':'12px';}
.form-group{'margin-bottom':'15px';}
.form-group label{'display':'block';'margin-bottom':'5px';'font-weight':'bold';}
.form-group input,.form-group textarea,.form-group select{'width':'100%';'padding':'10px';'border':'1px solid var(--border)';'border-radius':'8px';'background':'var(--bg)';'color':'var(--text)';}
.chat-messages{'height':'400px';'overflow-y':'auto';'border':'1px solid var(--border)';'border-radius':'12px';'padding':'15px';'background':'var(--bg)';'margin-bottom':'15px';}
.message{'margin-bottom':'15px';'padding':'10px';'background':'var(--card-bg)';'border-radius':'10px';'box-shadow':'var(--shadow)';}
.message-user{'font-weight':'bold';'color':'var(--dark-blue)';}
.message-time{'font-size':'11px';'color':'#999';'margin-left':'10px';}
.message-header{'display':'flex';'align-items':'center';'justify-content':'space-between';'margin-bottom':'8px';}
.media-preview{'max-width':'200px';'max-height':'150px';'margin-top':'10px';'border-radius':'8px';}
.chat-input-area{'display':'flex';'gap':'10px';'flex-wrap':'wrap';}
.chat-input-area textarea{'flex':'1';'padding':'10px';'border':'1px solid var(--border)';'border-radius':'8px';'background':'var(--bg)';'color':'var(--text)';}
.file-inputs{'margin-top':'10px';}
.error{'color':'red';'background':'#ffe0e0';'padding':'10px';'border-radius':'8px';'margin-bottom':'15px';}
.success{'color':'green';'background':'#e0ffe0';'padding':'10px';'border-radius':'8px';'margin-bottom':'15px';}
.notification,.transaction-item,.user-item,.deleted-item{'background':'var(--card-bg)';'padding':'12px';'margin-bottom':'10px';'border-radius':'10px';'border-left':'4px solid var(--light-blue)';}
.deleted-item{'border-left-color':'#ff6b6b';'background':'#fff8f0';}
.info-section{'background':'var(--card-bg)';'border-radius':'24px';'padding':'30px';'margin-bottom':'30px';'box-shadow':'var(--shadow)';}
.info-section h2{'color':'var(--dark-blue)';'margin-bottom':'20px';}
.reviews-section{'margin-top':'15px';'padding-top':'15px';'border-top':'1px solid var(--border)';}
.review{'background':'var(--bg)';'padding':'10px';'margin-bottom':'10px';'border-radius':'8px';}
.stars{'color':'#ffc107';'font-size':'16px';}
.modal{'display':'none';'position':'fixed';'z-index':1000;'left':0;'top':0;'width':'100%';'height':'100%';'background':'rgba(0,0,0,0.9);'}
.modal-content{'margin':'auto';'display':'block';'max-width':'90%';'max-height':'90%';'margin-top':'50px';}
.close{'position':'absolute';'top':'20px';'right':'35px';'color':'white';'font-size':'40px';'font-weight':'bold';'cursor':'pointer';}
.toggle-theme-btn{'background':'transparent';'border':'1px solid var(--light-blue)';'border-radius':'50%';'width':'40px';'height':'40px';'font-size':'20px';'cursor':'pointer';'transition':'0.3s';}
footer{'background':'var(--header-bg)';'color':'var(--text-light)';'padding':'30px 0';'text-align':'center';'margin-top':'40px';}
@media (max-width:768px){'.header-main{'flex-direction':'column';'text-align':'center';}'nav ul{'justify-content':'center';}'.cards-grid{'grid-template-columns':'1fr';}'.search-bar{'flex-direction':'column';}}
</style>
</head>
<body data-theme="{theme}">
<script>
function toggleTheme(){{
    let current = document.body.getAttribute("data-theme");
    let newTheme = current === "dark" ? "light" : "dark";
    document.body.setAttribute("data-theme", newTheme);
    fetch("/set_theme", {{method:"POST",headers:{{"Content-Type":"application/x-www-form-urlencoded"}},body:"theme="+newTheme}});
}}
</script>
<header><div class="top-banner">🏡 Лянтор | Ваш надёжный партнёр в мире недвижимости</div>
<div class="container header-main"><div class="logo-area"><div class="logo-icon">🏢</div><div class="logo-text"><h1>РИЭЛТОРСКОЕ АГЕНТСТВО</h1><h2>BY ГАУН</h2><div class="slogan">МЕСТО, ГДЕ ВЫ СТАНЕТЕ СОБОЙ</div></div></div>
<nav><ul><li><a href="/">Главная</a></li><li><a href="/buy">🏠 Покупка</a></li><li><a href="/sell">📝 Продажа</a></li><li><a href="/compare">📊 Сравнение</a></li><li><a href="/favorites">❤️ Избранное</a></li><li><a href="/chat">💬 Чат</a></li><li><a href="/info">📊 Сведения</a></li><li><a href="/contacts">📞 Контакты</a></li></ul></nav>
<div class="user-info">{theme_toggle}{user_html}</div></div></header>
<main>{err_html}{suc_html}{content}</main>
<footer><div class="container"><p>© 2025 Риэлторское агентство by Гаун | Лянтор</p></div></footer>
<div id="imageModal" class="modal" onclick="closeModal()"><span class="close">&times;</span><img class="modal-content" id="modalImage"></div>
<script>function openModal(src){{document.getElementById("modalImage").src=src;document.getElementById("imageModal").style.display="block";}}function closeModal(){{document.getElementById("imageModal").style.display="none";}}</script>
</body>
</html>'''

def get_user_from_session(request: Request):
    session_user = request.cookies.get("user_session")
    if session_user:
        try:
            user_data = json.loads(session_user)
            return user_data
        except:
            return None
    return None

def set_user_session(response, user_data):
    response.set_cookie(key="user_session", value=json.dumps(user_data), max_age=86400, httponly=True)

# ============= МАРШРУТЫ =============

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    user = get_user_from_session(request)
    if user and is_user_banned(user["username"]):
        response = RedirectResponse("/logout", 303)
        response.delete_cookie("user_session")
        return response
    
    props = get_all_properties()
    # Сортируем по просмотрам для главной
    props.sort(key=lambda x: x.get("views", 0), reverse=True)
    cards = ""
    for p in props[:6]:
        photos_html = ""
        if p.get("photos") and len(p["photos"]) > 0:
            photos_html = f'<div class="property-photos"><img src="{p["photos"][0]}" class="property-photo" onclick="openModal(this.src)"></div>'
        cards += f'<div class="property-card"><h3>{p["title"]}</h3><p><strong>📍 Локация:</strong> {p["location"]}</p><p><strong>💰 Цена:</strong> {p["price"]} ₽</p><p><strong>📐 Площадь:</strong> {p["area"]} м²</p><p><strong>👤 Продавец:</strong> {p["seller_name"]}</p>{photos_html}'
        if user and user.get("role") == "admin":
            cards += f'<form action="/delete_property/{p["id"]}" method="post"><button type="submit" class="delete-btn" onclick="return confirm(\'Удалить объявление?\')">🗑️ Удалить</button></form>'
        cards += '</div>'
    if not props:
        cards = '<p>Пока нет объявлений. Станьте продавцом и добавьте первое!</p>'
    
    content = f'''<section class="hero"><div class="container"><h1>Ваша недвижимость — наша профессиональная забота</h1><p>Поможем купить, продать или выкупить жильё в Лянторе</p></div></section>
    <section class="services-block"><div class="container"><h2 style="text-align:center;">Наши услуги</h2><p style="text-align:center;">Риэлторская компания / Наши услуги</p>
    <div class="services-grid"><div class="service-box"><h3>🏠 Риэлторские услуги</h3><ul><li>Купить недвижимость</li><li>Продать недвижимость</li><li>Обмен квартир в Лянторе</li><li>Анализ стоимости объекта</li><li>Срочный выкуп недвижимости</li><li>Помощь в получении кредита</li></ul><a href="/buy" class="btn">Подробнее →</a></div>
    <div class="service-box"><h3>⚡ Срочный выкуп недвижимости</h3><p><strong>Агентство «by Гаун» выкупит Вашу:</strong></p><ul><li>комнату, долю в квартире</li><li>квартиру (даже требующую ремонта)</li><li>новостройку на любой стадии</li></ul><a href="/sell" class="btn">Подробнее →</a></div></div></div></section>
    <section class="about-section"><div class="container"><h2>Риэлторское агентство</h2><p><strong>В чем суть работы риэлторского агентства?</strong> Это помощь клиенту в подборе и продаже недвижимости. Если вы не сталкиваетесь с рынком недвижимости каждый день, то вам довольно сложно ориентироваться в адекватности цен, разнообразии планировок и даже в удачности месторасположения объекта.</p>
    <h3>Агентство недвижимости г. Лянтора</h3><p>Работа риэлторского агентства заточена под то, чтобы помочь клиенту купить или продать свою недвижимость и сделать это максимально выгодно и комфортно.</p>
    <p>Чтобы любая сделка прошла безопасно на всех этапах, при выборе риэлторского агентства нужно учитывать такой важный фактор как репутация компании. Обращаясь в Агентство «by Гаун», можно быть уверенным в том, что ваши задачи будут решать профессионалы рынка.</p>
    <h3>Услуги риэлторского агентства в Лянторе</h3><ul class="about-list"><li><strong>Качественный подбор недвижимости</strong> от квалифицированных специалистов</li><li><strong>Быстрая и выгодная продажа объекта</strong> по максимальной цене</li><li><strong>Юридические услуги</strong> – сопровождение сделки и контроль документов</li></ul></div></section>
    <section style="padding:50px 0;"><div class="container"><h2 style="text-align:center;">Популярные объявления</h2><div class="cards-grid">{cards}</div></div></section>'''
    theme = user.get("theme", "light") if user else "light"
    return HTMLResponse(render_html(content, user=user, theme=theme))

@app.post("/set_theme")
async def set_theme(request: Request, theme: str = Form(...)):
    user = get_user_from_session(request)
    if user:
        update_user_by_username(user["username"], {"theme": theme})
        response = RedirectResponse(request.headers.get("referer", "/"), 303)
        set_user_session(response, {**user, "theme": theme})
        return response
    return RedirectResponse("/", 303)

@app.get("/buy", response_class=HTMLResponse)
async def buy(request: Request, search: str = "", min_price: str = "", max_price: str = "", min_area: str = "", max_area: str = "", type_filter: str = "", sort: str = ""):
    user = get_user_from_session(request)
    if user and is_user_banned(user["username"]):
        response = RedirectResponse("/logout", 303)
        response.delete_cookie("user_session")
        return response
    
    props = get_all_properties()
    
    # Фильтрация
    filtered = []
    for p in props:
        if search and search.lower() not in p["title"].lower() and search.lower() not in p["location"].lower() and search.lower() not in p["description"].lower():
            continue
        if min_price and p["price"] < int(min_price):
            continue
        if max_price and p["price"] > int(max_price):
            continue
        if min_area and p["area"] < float(min_area):
            continue
        if max_area and p["area"] > float(max_area):
            continue
        if type_filter and type_filter != "all":
            if type_filter == "sale" and p.get("listing_type") == "rent":
                continue
            if type_filter == "rent" and p.get("listing_type") == "sale":
                continue
        filtered.append(p)
    
    # Сортировка
    if sort == "price_asc":
        filtered.sort(key=lambda x: x["price"])
    elif sort == "price_desc":
        filtered.sort(key=lambda x: x["price"], reverse=True)
    elif sort == "date_desc":
        filtered.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    else:
        filtered.sort(key=lambda x: x.get("views", 0), reverse=True)
    
    user_favorites = []
    if user:
        user_data = get_user_by_username(user["username"])
        user_favorites = get_user_favorites(user_data["id"])
    
    cards = ""
    for p in filtered:
        is_fav = p["id"] in user_favorites
        fav_icon = "❤️" if is_fav else "🤍"
        fav_action = f'/remove_favorite/{p["id"]}' if is_fav else f'/add_favorite/{p["id"]}'
        
        photos_html = ""
        if p.get("photos") and len(p["photos"]) > 0:
            photos_html = '<div class="property-photos">'
            for photo in p["photos"][:3]:
                photos_html += f'<img src="{photo}" class="property-photo" onclick="openModal(this.src)">'
            photos_html += '</div>'
        
        reviews = get_reviews_by_property(p["id"])
        avg_rating = 0
        if reviews:
            avg_rating = sum(r["rating"] for r in reviews) / len(reviews)
        stars_html = "⭐" * round(avg_rating) + "☆" * (5 - round(avg_rating)) if avg_rating > 0 else "Нет отзывов"
        
        cards += f'<div class="property-card"><h3>{p["title"]}</h3><p><strong>📍 Локация:</strong> {p["location"]}</p><p><strong>💰 Цена:</strong> {p["price"]} ₽</p><p><strong>📐 Площадь:</strong> {p["area"]} м²</p><p><strong>🏗️ Этаж:</strong> {p["floor"]} / {p["total_floors"]}</p><p><strong>👀 Просмотров:</strong> {p.get("views", 0)}</p><p><strong>⭐ Рейтинг:</strong> {stars_html}</p>{photos_html}<div class="property-actions">'
        
        if user:
            cards += f'<form action="/add_to_comparison/{p["id"]}" method="post" style="display:inline;"><button type="submit" class="compare-btn" title="Добавить к сравнению">📊 Сравнить</button></form>'
            cards += f'<form action="{fav_action}" method="post" style="display:inline;"><button type="submit" class="fav-btn">{fav_icon}</button></form>'
            if user.get("role") in ["client", "seller", "admin"]:
                cards += f'<form action="/buy_property/{p["id"]}" method="post" style="display:inline;"><button type="submit" class="buy-btn" onclick="return confirm(\'Купить {p["title"]} за {p["price"]} ₽?\')">💰 Купить</button></form>'
                cards += f'<form action="/rent_property/{p["id"]}" method="post" style="display:inline;"><button type="submit" class="rent-btn" onclick="return confirm(\'Арендовать {p["title"]}?\')">🔑 Арендовать</button></form>'
            if user.get("role") == "admin":
                cards += f'<form action="/delete_property/{p["id"]}" method="post" style="display:inline;"><button type="submit" class="delete-btn" onclick="return confirm(\'Удалить объявление?\')">🗑️ Удалить</button></form>'
        
        cards += f'</div>'
        
        if user and user.get("role") == "client":
            cards += f'<details><summary>⭐ Оставить отзыв</summary><form action="/add_review/{p["id"]}" method="post" style="margin-top:10px;"><select name="rating" required><option value="5">5 ★ - Отлично</option><option value="4">4 ★ - Хорошо</option><option value="3">3 ★ - Нормально</option><option value="2">2 ★ - Плохо</option><option value="1">1 ★ - Ужасно</option></select><textarea name="comment" placeholder="Ваш отзыв (до 200 символов)" maxlength="200" rows="2" style="width:100%;margin-top:5px;"></textarea><button type="submit" class="btn" style="margin-top:5px;">Отправить</button></form></details>'
        
        cards += '<div class="reviews-section"><strong>⭐ Последние отзывы:</strong>'
        for r in reviews[-3:]:
            stars = '★' * r["rating"] + '☆' * (5 - r["rating"])
            cards += f'<div class="review"><span class="stars">{stars}</span> <strong>{r["nickname"]}:</strong> {r["comment"][:100]}</div>'
        if not reviews:
            cards += '<p>Нет отзывов. Будьте первым!</p>'
        cards += '</div></div>'
    
    if not filtered:
        cards = '<p>Нет объявлений, соответствующих критериям.</p>'
    
    content = f'''
    <section style="padding:50px 0;"><div class="container">
        <h1 style="color:var(--dark-blue);margin-bottom:30px;">🏠 Квартиры и дома на продажу/аренду</h1>
        <form method="get" class="search-bar">
            <div class="search-group"><label>🔍 Поиск</label><input type="text" name="search" placeholder="Название, адрес..." value="{search}"></div>
            <div class="search-group"><label>💰 Цена от</label><input type="number" name="min_price" placeholder="от" value="{min_price}"></div>
            <div class="search-group"><label>💰 Цена до</label><input type="number" name="max_price" placeholder="до" value="{max_price}"></div>
            <div class="search-group"><label>📐 Площадь от</label><input type="number" step="0.1" name="min_area" placeholder="от" value="{min_area}"></div>
            <div class="search-group"><label>📐 Площадь до</label><input type="number" step="0.1" name="max_area" placeholder="до" value="{max_area}"></div>
            <div class="search-group"><label>📋 Тип</label><select name="type_filter"><option value="all">Все</option><option value="sale" {'selected' if type_filter=="sale" else ''}>Продажа</option><option value="rent" {'selected' if type_filter=="rent" else ''}>Аренда</option></select></div>
            <div class="search-group"><label>📊 Сортировка</label><select name="sort"><option value="views" {'selected' if sort=="" or sort=="views" else ''}>По просмотрам</option><option value="price_asc" {'selected' if sort=="price_asc" else ''}>Цена (сначала дешёвые)</option><option value="price_desc" {'selected' if sort=="price_desc" else ''}>Цена (сначала дорогие)</option><option value="date_desc" {'selected' if sort=="date_desc" else ''}>Новые сначала</option></select></div>
            <div class="search-group"><button type="submit" class="btn" style="margin-top:23px;">🔍 Найти</button></div>
        </form>
        <div class="cards-grid">{cards}</div>
    </div></section>'''
    theme = user.get("theme", "light") if user else "light"
    return HTMLResponse(render_html(content, user=user, theme=theme))

@app.post("/add_favorite/{prop_id}")
async def add_favorite_route(request: Request, prop_id: int):
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse("/login?error=Авторизуйтесь", 303)
    user_data = get_user_by_username(user["username"])
    add_favorite(user_data["id"], prop_id)
    return RedirectResponse(request.headers.get("referer", "/buy"), 303)

@app.post("/remove_favorite/{prop_id}")
async def remove_favorite_route(request: Request, prop_id: int):
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse("/login?error=Авторизуйтесь", 303)
    user_data = get_user_by_username(user["username"])
    remove_favorite(user_data["id"], prop_id)
    return RedirectResponse(request.headers.get("referer", "/buy"), 303)

@app.get("/favorites", response_class=HTMLResponse)
async def favorites(request: Request):
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse("/login?error=Авторизуйтесь", 303)
    if is_user_banned(user["username"]):
        response = RedirectResponse("/logout", 303)
        response.delete_cookie("user_session")
        return response
    
    user_data = get_user_by_username(user["username"])
    fav_ids = get_user_favorites(user_data["id"])
    all_props = get_all_properties()
    fav_props = [p for p in all_props if p["id"] in fav_ids]
    
    cards = ""
    for p in fav_props:
        photos_html = ""
        if p.get("photos") and len(p["photos"]) > 0:
            photos_html = f'<div class="property-photos"><img src="{p["photos"][0]}" class="property-photo" onclick="openModal(this.src)"></div>'
        cards += f'<div class="property-card"><h3>{p["title"]}</h3><p><strong>📍 Локация:</strong> {p["location"]}</p><p><strong>💰 Цена:</strong> {p["price"]} ₽</p>{photos_html}<div class="property-actions"><form action="/remove_favorite/{p["id"]}" method="post"><button type="submit" class="delete-btn">❤️ Удалить из избранного</button></form><a href="/buy?search={p["title"]}" class="btn">🔍 Посмотреть</a></div></div>'
    if not cards:
        cards = "<p>У вас пока нет избранных объявлений. Добавьте их на странице поиска!</p>"
    
    content = f'<section style="padding:50px 0;"><div class="container"><h1 style="color:var(--dark-blue);">❤️ Моё избранное</h1><div class="cards-grid">{cards}</div></div></section>'
    theme = user.get("theme", "light")
    return HTMLResponse(render_html(content, user=user, theme=theme))
    @app.get("/compare", response_class=HTMLResponse)
async def compare(request: Request):
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse("/login?error=Авторизуйтесь", 303)
    
    user_data = get_user_by_username(user["username"])
    compare_ids = get_comparison(user_data["id"])
    all_props = get_all_properties()
    compare_props = [p for p in all_props if p["id"] in compare_ids]
    
    if not compare_props:
        content = '<section style="padding:50px 0;"><div class="container"><h1>📊 Сравнение объектов</h1><p>У вас нет объектов для сравнения. Добавьте их на странице поиска, нажав кнопку "Сравнить".</p><a href="/buy" class="btn">Перейти к поиску</a></div></section>'
        theme = user.get("theme", "light")
        return HTMLResponse(render_html(content, user=user, theme=theme))
    
    # Таблица сравнения
    table = '<table style="width:100%;border-collapse:collapse;background:var(--card-bg);border-radius:16px;overflow:hidden;"><tr style="background:var(--light-blue);"><th style="padding:12px;">Параметр</th>'
    for p in compare_props:
        table += f'<th style="padding:12px;">{p["title"]}</th>'
    table += '</tr>'
    
    rows = [
        ("📍 Локация", lambda x: x["location"]),
        ("💰 Цена", lambda x: f'{x["price"]} ₽'),
        ("📐 Площадь", lambda x: f'{x["area"]} м²'),
        ("🏗️ Этаж", lambda x: f'{x["floor"]} / {x["total_floors"]}'),
        ("🏠 Тип строения", lambda x: x["building_type"]),
        ("👀 Просмотров", lambda x: x.get("views", 0)),
        ("📝 Описание", lambda x: x["description"][:100] + "..."),
    ]
    
    for row_name, getter in rows:
        table += f'<tr style="border-bottom:1px solid var(--border);"><td style="padding:10px;font-weight:bold;">{row_name}</td>'
        for p in compare_props:
            table += f'<td style="padding:10px;">{getter(p)}</td>'
        table += '</tr>'
    table += '</table>'
    
    content = f'''
    <section style="padding:50px 0;"><div class="container">
        <h1 style="color:var(--dark-blue);">📊 Сравнение объектов</h1>
        {table}
        <div style="margin-top:20px;display:flex;gap:10px;">
            <form action="/clear_comparison" method="post"><button type="submit" class="delete-btn" onclick="return confirm('Очистить список сравнения?')">🗑️ Очистить сравнение</button></form>
            <a href="/buy" class="btn">➕ Добавить ещё</a>
        </div>
    </div></section>'''
    theme = user.get("theme", "light")
    return HTMLResponse(render_html(content, user=user, theme=theme))

@app.post("/add_to_comparison/{prop_id}")
async def add_to_comparison(request: Request, prop_id: int):
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse("/login?error=Авторизуйтесь", 303)
    user_data = get_user_by_username(user["username"])
    add_to_comparison(user_data["id"], prop_id)
    return RedirectResponse("/compare", 303)

@app.post("/clear_comparison")
async def clear_comparison_route(request: Request):
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse("/login?error=Авторизуйтесь", 303)
    user_data = get_user_by_username(user["username"])
    clear_comparison(user_data["id"])
    return RedirectResponse("/compare", 303)

@app.get("/sell", response_class=HTMLResponse)
async def sell(request: Request):
    user = get_user_from_session(request)
    if user and is_user_banned(user["username"]):
        response = RedirectResponse("/logout", 303)
        response.delete_cookie("user_session")
        return response
    
    if user and (user.get("role") == "seller" or user.get("role") == "admin"):
        content = '<section style="padding:50px 0;"><div class="container"><h1 style="color:var(--dark-blue);">📝 Добавить объявление</h1><form action="/add_property" method="post" enctype="multipart/form-data" style="max-width:600px;background:var(--card-bg);padding:30px;border-radius:20px;"><div class="form-group"><label>Название</label><input type="text" name="title" required></div><div class="form-group"><label>Локация</label><input type="text" name="location" required></div><div class="form-group"><label>Цена (₽)</label><input type="number" name="price" required></div><div class="form-group"><label>Площадь (м²)</label><input type="number" step="0.1" name="area" required></div><div class="form-group"><label>Этаж</label><input type="number" name="floor" required></div><div class="form-group"><label>Всего этажей</label><input type="number" name="total_floors" required></div><div class="form-group"><label>Тип строения</label><select name="building_type"><option>Кирпичный</option><option>Панельный</option><option>Монолитный</option><option>Деревянный</option><option>Блочный</option></select></div><div class="form-group"><label>Описание</label><textarea name="description" rows="4"></textarea></div><div class="form-group"><label>Фото (можно несколько, до 5 шт.)</label><input type="file" name="photos" multiple accept="image/*"></div><button type="submit" class="btn">Опубликовать</button></form></div></section>'
    else:
        content = '<section style="padding:50px 0;"><div class="container"><div style="background:var(--card-bg);padding:30px;border-radius:20px;text-align:center;"><p style="color:red;">⚠️ Только продавцы и администраторы могут добавлять объявления.</p><p><a href="/register">Зарегистрируйтесь как продавец</a></p></div></div></section>'
    theme = user.get("theme", "light") if user else "light"
    return HTMLResponse(render_html(content, user=user, theme=theme))

@app.post("/add_property")
async def add_property_route(request: Request, title: str = Form(...), location: str = Form(...), price: int = Form(...), area: float = Form(...), floor: int = Form(...), total_floors: int = Form(...), building_type: str = Form(...), description: str = Form(""), photos: List[UploadFile] = File(None)):
    user = get_user_from_session(request)
    if not user or user.get("role") not in ["seller", "admin"]:
        raise HTTPException(403, "Доступ запрещён")
    
    photo_paths = []
    if photos:
        for i, photo in enumerate(photos[:5]):
            if photo and photo.filename:
                ext = photo.filename.split(".")[-1]
                name = f"{uuid.uuid4()}.{ext}"
                path = f"static/uploads/property_photos/{name}"
                with open(path, "wb") as f:
                    shutil.copyfileobj(photo.file, f)
                photo_paths.append(f"/{path}")
    
    add_property(title, location, price, area, floor, total_floors, building_type, description, user.get("nickname", user["name"]), user["username"], photo_paths)
    
    # Уведомления для избранного
    all_users = get_all_users()
    for u in all_users:
        if u.get("role") == "client" and u.get("email") and u.get("email_verified"):
            # Можно добавить уведомление о новом объявлении
            add_notification(u["username"], f"🏠 Новое объявление: {title} за {price} ₽", datetime.now().strftime("%d.%m.%Y %H:%M"), "new_property", 0)
    
    return RedirectResponse("/buy", 303)

@app.post("/buy_property/{prop_id}")
async def buy_property(request: Request, prop_id: int):
    user = get_user_from_session(request)
    if not user or user.get("role") not in ["client", "seller", "admin"]:
        return RedirectResponse("/login?error=Авторизуйтесь для покупки", 303)
    
    props = get_all_properties()
    prop = next((p for p in props if p["id"] == prop_id), None)
    if not prop:
        return RedirectResponse("/buy?error=Объявление не найдено", 303)
    
    add_transaction(user["name"], prop["seller_name"], prop["title"], prop["price"], "Покупка", datetime.now().strftime("%d.%m.%Y %H:%M"))
    add_notification(prop["seller_username"], f"🏠 {user.get('nickname', user['name'])} КУПИЛ {prop['title']} за {prop['price']} ₽", datetime.now().strftime("%d.%m.%Y %H:%M"), "sale", prop_id)
    delete_property_by_id(prop_id)
    return RedirectResponse("/buy?success=Поздравляем с покупкой!", 303)

@app.post("/rent_property/{prop_id}")
async def rent_property(request: Request, prop_id: int):
    user = get_user_from_session(request)
    if not user or user.get("role") not in ["client", "seller", "admin"]:
        return RedirectResponse("/login?error=Авторизуйтесь для аренды", 303)
    
    props = get_all_properties()
    prop = next((p for p in props if p["id"] == prop_id), None)
    if not prop:
        return RedirectResponse("/buy?error=Объявление не найдено", 303)
    
    add_transaction(user["name"], prop["seller_name"], prop["title"], prop["price"], "Аренда", datetime.now().strftime("%d.%m.%Y %H:%M"))
    add_notification(prop["seller_username"], f"🔑 {user.get('nickname', user['name'])} АРЕНДОВАЛ {prop['title']}", datetime.now().strftime("%d.%m.%Y %H:%M"), "rent", prop_id)
    delete_property_by_id(prop_id)
    return RedirectResponse("/buy?success=Поздравляем с арендой!", 303)

@app.post("/add_review/{prop_id}")
async def add_review_route(request: Request, prop_id: int, rating: int = Form(...), comment: str = Form("")):
    user = get_user_from_session(request)
    if not user or user.get("role") != "client":
        return RedirectResponse("/login?error=Только клиенты могут оставлять отзывы", 303)
    
    add_review(prop_id, user["username"], user.get("nickname", user["name"]), rating, comment[:200])
    return RedirectResponse(f"/buy?success=Отзыв добавлен", 303)

@app.post("/delete_property/{prop_id}")
async def delete_property_route(request: Request, prop_id: int):
    user = get_user_from_session(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(403, "Только для администратора")
    delete_property_by_id(prop_id)
    return RedirectResponse("/buy", 303)

@app.get("/chat", response_class=HTMLResponse)
async def chat(request: Request):
    user = get_user_from_session(request)
    if user and is_user_banned(user["username"]):
        response = RedirectResponse("/logout", 303)
        response.delete_cookie("user_session")
        return response
    
    msgs = get_all_messages()
    html = ""
    for m in msgs[-50:]:
        btn = ''
        if user and (user.get("role") == "admin" or user.get("username") == m["username"]):
            btn = f'<form action="/delete_message/{m["id"]}" method="post" style="display:inline;"><button type="submit" class="delete-msg-btn" onclick="return confirm(\'Удалить сообщение?\')">🗑️</button></form>'
        html += f'<div class="message"><div class="message-header"><div><span class="message-user">{m["nickname"]}</span><span class="message-time">{m["time"]}</span></div><div>{btn}</div></div><p>{m["text"]}</p>'
        if m.get("photo"):
            html += f'<img src="{m["photo"]}" class="media-preview">'
        html += '</div>'
    if not msgs:
        html = '<p style="text-align:center;color:#999;">Пока нет сообщений</p>'
    
    form = '<form action="/send_message" method="post" enctype="multipart/form-data"><div class="chat-input-area"><textarea name="message_text" rows="2" placeholder="Ваше сообщение..." required></textarea></div><div class="file-inputs"><input type="file" name="photo" accept="image/*"></div><button type="submit" class="btn" style="margin-top:10px;">Отправить</button></form>' if user else '<p><a href="/login">Войдите</a>, чтобы писать в чат</p>'
    
    # Личные сообщения
    all_users = get_all_users()
    user_list = '<div style="margin-top:20px;"><h2>📨 Личные сообщения</h2><div style="display:flex;gap:10px;flex-wrap:wrap;">'
    for u in all_users:
        if u["username"] != user["username"]:
            unread = len([m for m in get_private_messages(user["username"], u["username"]) if m["to_user"] == user["username"] and not m["read"]])
            badge = f' <span style="background:red;color:white;border-radius:50%;padding:2px 6px;font-size:10px;">{unread}</span>' if unread > 0 else ''
            user_list += f'<a href="/private_chat/{u["username"]}" class="btn">{u.get("nickname", u["username"])}{badge}</a>'
    user_list += '</div></div>'
    
    content = f'<section style="padding:50px 0;"><div class="container"><h1>💬 Общая переписка</h1><div class="chat-messages" id="chatMessages">{html}</div>{form}{user_list}</div></section><script>function loadMessages(){{fetch("/get_messages").then(r=>r.json()).then(d=>{{const c=document.getElementById("chatMessages");if(d.html){{c.innerHTML=d.html;c.scrollTop=c.scrollHeight;}}}});}}setInterval(loadMessages,3000);loadMessages();</script>'
    theme = user.get("theme", "light") if user else "light"
    return HTMLResponse(render_html(content, user=user, theme=theme))

@app.get("/private_chat/{to_user}", response_class=HTMLResponse)
async def private_chat(request: Request, to_user: str):
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse("/login", 303)
    
    mark_private_messages_read(user["username"], to_user)
    msgs = get_private_messages(user["username"], to_user)
    to_user_data = get_user_by_username(to_user)
    to_nickname = to_user_data.get("nickname", to_user) if to_user_data else to_user
    
    msgs_html = ""
    for m in msgs:
        msgs_html += f'<div class="message"><div class="message-header"><div><span class="message-user">{m["from_nickname"]}</span><span class="message-time">{m["time"]}</span></div></div><p>{m["text"]}</p></div>'
    if not msgs:
        msgs_html = '<p>Сообщений пока нет. Напишите что-нибудь!</p>'
    
    content = f'''
    <section style="padding:50px 0;"><div class="container">
        <h1>💬 Чат с {to_nickname}</h1>
        <div class="chat-messages" id="chatMessages">{msgs_html}</div>
        <form action="/send_private_message/{to_user}" method="post">
            <div class="chat-input-area"><textarea name="message_text" rows="2" placeholder="Ваше сообщение..." required></textarea></div>
            <button type="submit" class="btn" style="margin-top:10px;">Отправить</button>
        </form>
        <a href="/chat" class="btn" style="margin-top:20px;">← Назад к чатам</a>
    </div></section>
    <script>
        function loadMessages(){{fetch("/get_private_messages/{to_user}").then(r=>r.json()).then(d=>{{const c=document.getElementById("chatMessages");if(d.html){{c.innerHTML=d.html;c.scrollTop=c.scrollHeight;}}}});}}
        setInterval(loadMessages,3000);
        loadMessages();
    </script>'''
    theme = user.get("theme", "light")
    return HTMLResponse(render_html(content, user=user, theme=theme))

@app.post("/send_private_message/{to_user}")
async def send_private_message(request: Request, to_user: str, message_text: str = Form(...)):
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse("/login", 303)
    add_private_message(user["username"], user.get("nickname", user["name"]), to_user, message_text, datetime.now().strftime("%H:%M:%S"))
    add_notification(to_user, f"💬 Новое личное сообщение от {user.get('nickname', user['name'])}", datetime.now().strftime("%d.%m.%Y %H:%M"), "private_message", 0)
    return RedirectResponse(f"/private_chat/{to_user}", 303)

@app.get("/get_private_messages/{with_user}")
async def get_private_messages_json(request: Request, with_user: str):
    user = get_user_from_session(request)
    if not user:
        return {"html": ""}
    msgs = get_private_messages(user["username"], with_user)
    html = ""
    for m in msgs:
        html += f'<div class="message"><div class="message-header"><div><span class="message-user">{m["from_nickname"]}</span><span class="message-time">{m["time"]}</span></div></div><p>{m["text"]}</p></div>'
    if not html:
        html = '<p>Сообщений пока нет. Напишите что-нибудь!</p>'
    return {"html": html}

@app.post("/send_message")
async def send_message_route(request: Request, message_text: str = Form(...), photo: Optional[UploadFile] = File(None)):
    user = get_user_from_session(request)
    if not user:
        raise HTTPException(403, "Авторизуйтесь")
    if is_user_banned(user["username"]):
        raise HTTPException(403, "Вы забанены")
    
    path = None
    if photo and photo.filename:
        ext = photo.filename.split(".")[-1]
        name = f"{uuid.uuid4()}.{ext}"
        path = f"static/uploads/photos/{name}"
        with open(path, "wb") as f:
            shutil.copyfileobj(photo.file, f)
        path = f"/{path}"
    add_message(user["username"], user.get("nickname", user["name"]), message_text, datetime.now().strftime("%H:%M:%S"), path)
    return RedirectResponse("/chat", 303)

@app.post("/delete_message/{msg_id}")
async def delete_message_route(request: Request, msg_id: int):
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse("/login?error=Авторизуйтесь", 303)
    
    msgs = get_all_messages()
    msg = next((m for m in msgs if m["id"] == msg_id), None)
    if not msg:
        return RedirectResponse("/chat?error=Сообщение не найдено", 303)
    if user.get("role") != "admin" and user.get("username") != msg["username"]:
        return RedirectResponse("/chat?error=Нельзя удалить чужое сообщение", 303)
    
    delete_message_by_id(msg_id, user.get("nickname", user["name"]))
    return RedirectResponse("/chat?success=Сообщение удалено", 303)

@app.get("/get_messages")
async def get_messages_json(request: Request):
    user = get_user_from_session(request)
    msgs = get_all_messages()
    html = ""
    for m in msgs[-50:]:
        btn = ''
        if user and (user.get("role") == "admin" or user.get("username") == m["username"]):
            btn = f'<form action="/delete_message/{m["id"]}" method="post" style="display:inline;"><button type="submit" class="delete-msg-btn" onclick="return confirm(\'Удалить сообщение?\')">🗑️</button></form>'
        html += f'<div class="message"><div class="message-header"><div><span class="message-user">{m["nickname"]}</span><span class="message-time">{m["time"]}</span></div><div>{btn}</div></div><p>{m["text"]}</p>'
        if m.get("photo"):
            html += f'<img src="{m["photo"]}" class="media-preview">'
        html += '</div>'
    if not msgs:
        html = '<p style="text-align:center;color:#999;">Пока нет сообщений</p>'
    return {"html": html}

@app.get("/info", response_class=HTMLResponse)
async def info(request: Request):
    user = get_user_from_session(request)
    if user and is_user_banned(user["username"]):
        response = RedirectResponse("/logout", 303)
        response.delete_cookie("user_session")
        return response
    
    trans = get_all_transactions()
    trans_html = ""
    for t in trans[:20]:
        trans_html += f'<div class="transaction-item"><strong>📅 {t["date"]}</strong><p>🏠 {t["property_title"]}</p><p>💰 {t["price"]} ₽</p><p>👤 Продавец: {t["seller"]} → Покупатель: {t["buyer"]}</p><p>📝 Тип: {t["type"]}</p></div>'
    if not trans:
        trans_html = "<p>Пока нет завершённых сделок</p>"
    
    admin_html = ""
    if user and user.get("role") == "admin":
        users_list = get_all_users()
        users_html = ""
        for u in users_list:
            ban_status = "🔴 Забанен" if u.get("is_banned") else "🟢 Активен"
            email_status = "✅ Подтверждён" if u.get("email_verified") else "❌ Не подтверждён"
            ban_button = ''
            if not u.get("is_banned") and u["username"] != "artem_gaun":
                ban_button = f'<form action="/ban_user/{u["username"]}" method="post" style="display:inline;"><button type="submit" class="ban-btn" onclick="return confirm(\'Забанить {u.get("nickname", u["name"])}?\')">🚫 Забанить</button></form>'
            elif u.get("is_banned"):
                ban_button = f'<form action="/unban_user/{u["username"]}" method="post" style="display:inline;"><button type="submit" class="unban-btn" onclick="return confirm(\'Разбанить {u.get("nickname", u["name"])}?\')">✅ Разбанить</button></form>'
            
            users_html += f'<div class="user-item"><strong>👤 {u.get("nickname", u["name"])}</strong> (@{u["username"]})<br>Роль: {u["role"]}<br>Почта: {u.get("email","не указана")} ({email_status})<br>Статус: {ban_status}<br>{ban_button}</div>'
        
        deleted = get_deleted_messages()
        del_html = ""
        for d in deleted[:30]:
            del_html += f'<div class="deleted-item"><strong>🗑️ {d["nickname"]}</strong> <small>{d["time"]}</small><p>{d["text"]}</p>'
            if d.get("photo"):
                del_html += f'<img src="{d["photo"]}" style="max-width:100px;border-radius:8px;">'
            del_html += f'<div style="font-size:12px;">Удалено: {d["deleted_by"]} ({d["deleted_at"]})</div></div>'
        
        props = get_all_properties()
        total_views = sum(p.get("views", 0) for p in props)
        
        # График продаж по месяцам
        monthly_sales = {}
        for t in trans:
            month = t["date"][:7] if len(t["date"]) > 7 else t["date"]
            monthly_sales[month] = monthly_sales.get(month, 0) + t["price"]
        
        chart_data = sorted(monthly_sales.items())[-6:]
        chart_labels = [c[0] for c in chart_data]
        chart_values = [c[1] for c in chart_data]
        
        admin_html = f'''
        <div class="info-section"><h2>📈 Статистика</h2>
            <p><strong>Всего объявлений:</strong> {len(props)}</p>
            <p><strong>Всего просмотров:</strong> {total_views}</p>
            <p><strong>Всего пользователей:</strong> {len(users_list)}</p>
            <p><strong>Завершённых сделок:</strong> {len(trans)}</p>
            <canvas id="salesChart" style="max-width:100%;height:300px;"></canvas>
        </div>
        <div class="info-section"><h2>👥 Зарегистрированные пользователи</h2><div>{users_html if users_html else "<p>Нет пользователей</p>"}</div></div>
        <div class="info-section"><h2>🗑️ Корзина</h2><div>{del_html if deleted else "<p>Корзина пуста</p>"}</div><form action="/clear_deleted_messages" method="post" style="margin-top:10px;"><button type="submit" class="delete-btn" onclick="return confirm(\'Очистить корзину полностью?\')">🗑️ Очистить корзину</button></form></div>
        <div class="info-section"><h2>📎 Экспорт отчётов</h2><a href="/export_transactions" class="btn">📊 Выгрузить сделки в Excel</a></div>
        <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
        <script>
            const ctx = document.getElementById('salesChart').getContext('2d');
            new Chart(ctx, {{
                type: 'bar',
                data: {{ labels: {chart_labels}, datasets: [{{ label: 'Выручка (₽)', data: {chart_values}, backgroundColor: '#6ec8e6' }}] }},
                options: {{ responsive: true, maintainAspectRatio: true }}
            }});
        </script>'''
    
    content = f'<section style="padding:50px 0;"><div class="container"><h1 style="color:var(--dark-blue);">📊 Сведения о сделках</h1><div class="info-section"><h2>📋 История покупок и продаж</h2><div>{trans_html}</div></div>{admin_html}</div></section>'
    theme = user.get("theme", "light") if user else "light"
    return HTMLResponse(render_html(content, user=user, theme=theme))

@app.get("/export_transactions")
async def export_transactions(request: Request):
    user = get_user_from_session(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(403, "Только для администратора")
    
    trans = get_all_transactions()
    wb = export_transactions_to_excel(trans)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(content=buffer.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=transactions.xlsx"})

@app.post("/clear_deleted_messages")
async def clear_deleted_messages_route(request: Request):
    user = get_user_from_session(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(403, "Только для администратора")
    clear_deleted_messages()
    return RedirectResponse("/info?success=Корзина очищена", 303)

@app.post("/ban_user/{username}")
async def ban_user_route(request: Request, username: str):
    user = get_user_from_session(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(403, "Только для администратора")
    if username == "artem_gaun":
        return RedirectResponse("/info?error=Нельзя забанить главного администратора", 303)
    ban_user(username, user.get("nickname", user["name"]))
    return RedirectResponse("/info?success=Пользователь забанен", 303)

@app.post("/unban_user/{username}")
async def unban_user_route(request: Request, username: str):
    user = get_user_from_session(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(403, "Только для администратора")
    unban_user(username)
    return RedirectResponse("/info?success=Пользователь разбанен", 303)

@app.get("/contacts", response_class=HTMLResponse)
async def contacts(request: Request):
    user = get_user_from_session(request)
    if user and is_user_banned(user["username"]):
        response = RedirectResponse("/logout", 303)
        response.delete_cookie("user_session")
        return response
    
    content = '<section style="padding:50px 0;"><div class="container"><div style="background:var(--card-bg);border-radius:24px;padding:40px;"><h1 style="color:var(--dark-blue);">📬 Наши контакты</h1><div style="background:var(--bg);padding:22px;border-radius:16px;margin:25px 0;"><p><strong>📍 Адрес:</strong> г. Лянтор, Лянторский Нефтяной техникум</p><p><strong>📧 Email:</strong> artemgaun104@gmail.com</p></div><div style="border-radius:20px;overflow:hidden;"><iframe src="https://yandex.ru/map-widget/v1/?ll=72.1579%2C61.6229&z=17&pt=72.1589,61.6230&what=here%3A1&lang=ru_RU" style="width:100%;height:380px;border:0;"></iframe></div></div></div></section>'
    theme = user.get("theme", "light") if user else "light"
    return HTMLResponse(render_html(content, user=user, theme=theme))

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request, error: str = None):
    user = get_user_from_session(request)
    if user:
        return RedirectResponse("/", 303)
    
    err = f'<div class="error">{error}</div>' if error else ''
    content = f'<section style="padding:50px 0;"><div class="container"><div style="max-width:400px;margin:0 auto;background:var(--card-bg);padding:30px;border-radius:20px;"><h2 style="color:var(--dark-blue);">Вход</h2>{err}<form action="/login" method="post"><div class="form-group"><label>Логин</label><input type="text" name="username" required></div><div class="form-group"><label>Пароль</label><input type="password" name="password" required></div><button type="submit" class="btn">Войти</button></form><p><a href="/forgot_password">Забыли пароль?</a></p><p><a href="/register">Зарегистрироваться</a></p></div></div></section>'
    return HTMLResponse(render_html(content, user=user))

@app.post("/login")
async def login_post(request: Request, username: str = Form(...), password: str = Form(...)):
    user = get_user_by_username(username)
    if user and user["password"] == password and user["is_active"]:
        if is_user_banned(username):
            return RedirectResponse("/login?error=Ваш аккаунт заблокирован", 303)
        if user.get("email") and not user.get("email_verified"):
            return RedirectResponse("/login?error=Подтвердите email. Проверьте почту!", 303)
        response = RedirectResponse("/", 303)
        set_user_session(response, {"username": user["username"], "role": user["role"], "name": user["name"], "nickname": user.get("nickname", user["name"]), "theme": user.get("theme", "light")})
        # Запись просмотров будет при открытии объявлений
        return response
    return RedirectResponse("/login?error=Неверный логин или пароль", 303)

@app.get("/logout")
async def logout():
    response = RedirectResponse("/", 303)
    response.delete_cookie("user_session")
    return response

@app.get("/register", response_class=HTMLResponse)
async def register_page(request: Request, error: str = None):
    user = get_user_from_session(request)
    if user:
        return RedirectResponse("/", 303)
    
    err = f'<div class="error">{error}</div>' if error else ''
    content = f'<section style="padding:50px 0;"><div class="container"><div style="max-width:400px;margin:0 auto;background:var(--card-bg);padding:30px;border-radius:20px;"><h2 style="color:var(--dark-blue);">Регистрация</h2>{err}<form action="/register" method="post"><div class="form-group"><label>Имя</label><input type="text" name="name" required></div><div class="form-group"><label>Логин</label><input type="text" name="username" required></div><div class="form-group"><label>Никнейм (будет виден в чате)</label><input type="text" name="nickname" required></div><div class="form-group"><label>Email</label><input type="email" name="email" required></div><div class="form-group"><label>Пароль</label><input type="password" name="password" required></div><div class="form-group"><label>Роль</label><select name="role"><option value="client">Клиент</option><option value="seller">Продавец</option></select></div><button type="submit" class="btn">Зарегистрироваться</button></form><p><a href="/login">Уже есть аккаунт?</a></p></div></div></section>'
    return HTMLResponse(render_html(content, user=user))

@app.post("/register")
async def register_post(request: Request, name: str = Form(...), username: str = Form(...), nickname: str = Form(...), email: str = Form(...), password: str = Form(...), role: str = Form(...)):
    if get_user_by_username(username):
        return RedirectResponse("/register?error=Пользователь уже существует", 303)
    if get_user_by_email(email):
        return RedirectResponse("/register?error=Email уже зарегистрирован", 303)
    if get_user_by_nickname(nickname):
        return RedirectResponse("/register?error=Никнейм уже занят", 303)
    if not is_valid_email(email):
        return RedirectResponse("/register?error=Неверный формат почты. Поддерживается @mail.ru или @gmail.com", 303)
    
    create_user(username, password, nickname, name, role, email)
    user = get_user_by_username(username)
    code = create_verification_code(user["id"], email)
    send_verification_email(email, nickname, code)
    return RedirectResponse("/login?success=Регистрация успешна! Подтвердите email, перейдя по ссылке в письме", 303)

@app.get("/verify_email")
async def verify_email(request: Request, code: str):
    if verify_email_code(code):
        return RedirectResponse("/login?success=Email подтверждён! Теперь вы можете войти", 303)
    return RedirectResponse("/login?error=Неверный или просроченный код подтверждения", 303)

@app.get("/forgot_password", response_class=HTMLResponse)
async def forgot_password_page(request: Request, error: str = None):
    user = get_user_from_session(request)
    err = f'<div class="error">{error}</div>' if error else ''
    content = f'<section style="padding:50px 0;"><div class="container"><div style="max-width:400px;margin:0 auto;background:var(--card-bg);padding:30px;border-radius:20px;"><h2 style="color:var(--dark-blue);">Восстановление пароля</h2>{err}<form action="/forgot_password" method="post"><div class="form-group"><label>Email</label><input type="email" name="email" required></div><button type="submit" class="btn">Отправить ссылку</button></form><p><a href="/login">Вспомнили пароль?</a></p></div></div></section>'
    return HTMLResponse(render_html(content, user=user))

@app.post("/forgot_password")
async def forgot_password_post(email: str = Form(...)):
    user = get_user_by_email(email)
    if user:
        code = create_reset_code(user["id"], email)
        send_reset_email(email, user.get("nickname", user["name"]), code)
    return RedirectResponse("/login?success=Если email зарегистрирован, вы получите письмо с инструкциями", 303)

@app.get("/reset_password_form")
async def reset_password_form(request: Request, code: str, error: str = None):
    user_id, email = verify_reset_code(code)
    if not user_id:
        return RedirectResponse("/login?error=Неверная или просроченная ссылка", 303)
    
    err = f'<div class="error">{error}</div>' if error else ''
    content = f'<section style="padding:50px 0;"><div class="container"><div style="max-width:400px;margin:0 auto;background:var(--card-bg);padding:30px;border-radius:20px;"><h2 style="color:var(--dark-blue);">Новый пароль</h2>{err}<form action="/reset_password" method="post"><input type="hidden" name="code" value="{code}"><div class="form-group"><label>Новый пароль</label><input type="password" name="new_password" required></div><div class="form-group"><label>Подтверждение</label><input type="password" name="confirm_password" required></div><button type="submit" class="btn">Сохранить</button></form></div></div></section>'
    return HTMLResponse(render_html(content, user=None))

@app.post("/reset_password")
async def reset_password_post(code: str = Form(...), new_password: str = Form(...), confirm_password: str = Form(...)):
    if new_password != confirm_password:
        return RedirectResponse(f"/reset_password_form?code={code}&error=Пароли не совпадают", 303)
    user_id, email = verify_reset_code(code)
    if not user_id:
        return RedirectResponse("/login?error=Неверная или просроченная ссылка", 303)
    
    user = get_user_by_username(get_user_by_email(email)["username"])
    if user:
        update_user_by_username(user["username"], {"password": new_password})
        delete_reset_code(code)
    return RedirectResponse("/login?success=Пароль успешно изменён. Теперь вы можете войти", 303)

@app.get("/profile", response_class=HTMLResponse)
async def profile(request: Request, error: str = None, success: str = None):
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse("/login", 303)
    if is_user_banned(user["username"]):
        response = RedirectResponse("/logout", 303)
        response.delete_cookie("user_session")
        return response
    
    mark_notifications_read(user["username"])
    notif_list = get_notifications_by_user(user["username"])
    notif_html = ""
    for n in notif_list:
        notif_html += f'<div class="notification"><strong>{n["date"]}</strong><p>{n["message"]}</p></div>'
    if not notif_html:
        notif_html = "<p>У вас пока нет уведомлений</p>"
    
    user_data = get_user_by_username(user["username"])
    email = user_data.get("email", "") if user_data else ""
    nickname = user_data.get("nickname", "") if user_data else ""
    email_verified = user_data.get("email_verified", False)
    email_status = "✅ Подтверждён" if email_verified else "❌ Не подтверждён"
    
    # История просмотров
    user_full = get_user_by_username(user["username"])
    history = get_user_view_history(user_full["id"])
    all_props = get_all_properties()
    history_props = []
    for h in history:
        for p in all_props:
            if p["id"] == h["property_id"]:
                history_props.append(p)
                break
    
    history_html = ""
    for p in history_props[:10]:
        history_html += f'<div class="property-card" style="margin-bottom:10px;"><a href="/buy?search={p["title"]}" style="text-decoration:none;color:inherit;"><h4>{p["title"]} - {p["price"]} ₽</h4></a></div>'
    if not history_html:
        history_html = "<p>Вы ещё не просматривали объявления</p>"
    
    content = f'''
    <section style="padding:50px 0;"><div class="container">
        <div style="background:var(--card-bg);border-radius:24px;padding:40px;">
            <h1 style="color:var(--dark-blue);">📋 Личный кабинет</h1>
            <div style="margin:30px 0;padding:20px;background:var(--bg);border-radius:16px;">
                <p><strong>👤 Имя:</strong> {user["name"]}</p>
                <p><strong>🔑 Никнейм:</strong> {nickname}</p>
                <p><strong>🎭 Роль:</strong> {user["role"]}</p>
                <p><strong>📧 Email:</strong> {email} ({email_status})</p>
            </div>
            <form action="/update_nickname" method="post"><div class="form-group"><label>Изменить никнейм</label><div style="display:flex;gap:10px;"><input type="text" name="nickname" placeholder="Новый никнейм" value="{nickname}" style="flex:1;" required><button type="submit" class="btn">Обновить</button></div></div></form>
            <form action="/update_email" method="post"><div class="form-group"><label>Изменить почту</label><div style="display:flex;gap:10px;"><input type="email" name="email" placeholder="example@mail.ru" value="{email}" style="flex:1;"><button type="submit" class="btn">Сохранить</button></div></div></form>
            {'' if email_verified else '<div style="margin:15px 0;"><a href="/send_verification" class="btn">📧 Отправить повторно письмо для подтверждения</a></div>'}
            <h2 style="color:var(--dark-blue);margin:30px 0 20px;">🔔 Уведомления</h2>
            <div>{notif_html}</div>
            <h2 style="color:var(--dark-blue);margin:30px 0 20px;">🕐 История просмотров</h2>
            <div>{history_html}</div>
        </div>
    </div></section>'''
    theme = user.get("theme", "light")
    return HTMLResponse(render_html(content, user=user, theme=theme))

@app.get("/send_verification")
async def send_verification(request: Request):
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse("/login", 303)
    user_data = get_user_by_username(user["username"])
    if user_data.get("email_verified"):
        return RedirectResponse("/profile?error=Email уже подтверждён", 303)
    code = create_verification_code(user_data["id"], user_data["email"])
    send_verification_email(user_data["email"], user_data.get("nickname", user_data["name"]), code)
    return RedirectResponse("/profile?success=Письмо отправлено повторно", 303)

@app.post("/update_nickname")
async def update_nickname(request: Request, nickname: str = Form(...)):
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse("/login", 303)
    if get_user_by_nickname(nickname) and get_user_by_nickname(nickname)["username"] != user["username"]:
        return RedirectResponse("/profile?error=Никнейм уже занят", 303)
    update_user_by_username(user["username"], {"nickname": nickname})
    response = RedirectResponse("/profile?success=Никнейм обновлён", 303)
    set_user_session(response, {**user, "nickname": nickname})
    return response

@app.post("/update_email")
async def update_email(request: Request, email: str = Form(...)):
    user = get_user_from_session(request)
    if not user:
        return RedirectResponse("/login", 303)
    if email and not is_valid_email(email):
        return RedirectResponse("/profile?error=Неверный формат почты. Поддерживается @mail.ru или @gmail.com", 303)
    if get_user_by_email(email) and get_user_by_email(email)["username"] != user["username"]:
        return RedirectResponse("/profile?error=Email уже используется другим пользователем", 303)
    update_user_by_username(user["username"], {"email": email, "email_verified": False})
    user_data = get_user_by_username(user["username"])
    code = create_verification_code(user_data["id"], email)
    send_verification_email(email, user.get("nickname", user["name"]), code)
    return RedirectResponse("/profile?success=Почта обновлена. Подтвердите её, перейдя по ссылке в письме", 303)

@app.middleware("http")
async def track_views(request: Request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/buy_property") or request.url.path.startswith("/buy") or request.url.path.startswith("/rent_property"):
        user = get_user_from_session(request)
        if user and request.url.path.startswith("/buy"):
            # Запись просмотра при открытии страницы покупки не делаем, только при прямом просмотре объявления
            pass
    return response

# Просмотр объявления с записью
@app.get("/view_property/{prop_id}")
async def view_property(request: Request, prop_id: int):
    user = get_user_from_session(request)
    increment_property_views(prop_id)
    if user:
        user_data = get_user_by_username(user["username"])
        add_view_history(user_data["id"], prop_id)
    return RedirectResponse(f"/buy?search={prop_id}", 303)

if __name__ == "__main__":
    print("\n" + "="*60)
    print("✅ САЙТ ЗАПУЩЕН С ПОЛНЫМ ФУНКЦИОНАЛОМ!")
    print("🌐 http://127.0.0.1:8000")
    print("="*60)
    print("\n🔑 АДМИНИСТРАТОР:")
    print("   Логин: artem_gaun")
    print("   Пароль: Admin_321")
    print("\n🎯 ВСЕ НОВЫЕ ФУНКЦИИ АКТИВИРОВАНЫ:")
    print("   • Поиск и фильтрация")
    print("   • Панель админа с графиками")
    print("   • Избранное")
    print("   • Умные уведомления на почту")
    print("   • История просмотров")
    print("   • Сравнение объектов")
    print("   • Личные сообщения")
    print("   • Счётчик просмотров")
    print("   • Восстановление пароля")
    print("   • Экспорт в Excel")
    print("   • Подтверждение email")
    print("   • Тёмная тема")
    print("="*60 + "\n")
    uvicorn.run(app, host="0.0.0.0", port=8000)
